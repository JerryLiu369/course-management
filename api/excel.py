import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Side, Border
from map import *

'''
map 获取具体部类核心课，专业核心课
excel 制作excel表格
'''


# 创建一个excel表格对象
wb = openpyxl.Workbook()
# 获取当前活跃的sheet页,默认就是第一个sheet页
ws = wb.active
align = Alignment(horizontal='center', vertical='center', text_rotation=0, 
                  wrap_text=True, shrink_to_fit=True, indent=1)
side = Side(style="medium", color="FFFFFF")

#major_name 需要从前端选择，然后获取
major_name = '计算机科学与技术'
max1,ls_buleihexin = map_bulei(major_name)
max2,ls_zhuanyehexin = map_zhuan(major_name)
# max1 = 3 #部类核心
# max2 = 4 #专业核心
# ls_buleihexin = [['高等数学','人工智能与Python程序设计','c'],['d','e'],['f'],[''],[''],[''],['g','h'],['i']]
# ls_zhuanyehexin = [['aa','bb','cc','q'],['d','e'],['f'],[''],[''],[''],['g','h'],['i']]

#赋值
major_cell = ws.cell(row=1, column=1)
major_cell.value = major_name
#对齐
major_cell.alignment = align
#字体
major_cell.font = Font(name="微软雅黑", size=20, color='F8F8FF', bold=True, 
            italic=False, strike=False)
#填充
major_cell.fill = PatternFill(fill_type='solid', start_color="4682B4", end_color='00FF00')
#边框
major_cell.border = Border(bottom=side)
#合并
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

#课程模块
course_category = ws.cell(row=2, column=1)
course_category.value = '课程模块'
course_category.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, 
                  wrap_text=False, shrink_to_fit=False, indent=0)
course_category.font = Font(name="宋体 ", size=15, color='000000', bold=True, 
            italic=False, strike=False)
course_category.fill = PatternFill(fill_type='solid', start_color="B0C4DE", end_color='00FF00')

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

#学期
for i in range(3,11):
    string = '第'+str(i-2) + '学期'
    term_cell = ws.cell(row=2, column=i)
    term_cell.value = string
    term_cell.alignment = align
    term_cell.font = Font(name="宋体 ", size=15, color='000000', bold=True, 
            italic=False, strike=False)
    term_cell.fill = PatternFill(fill_type='solid', start_color="B0C4DE", end_color='00FF00')
    term_cell.border = Border(left=side)

for i in range(1,6+max1+max2+1):
    ws.row_dimensions[i].height = 45

ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 11
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 22
ws.column_dimensions['I'].width = 22
ws.column_dimensions['J'].width = 22

#专业培养
peiyang_cell = ws.cell(row=3, column=1)
peiyang_cell.value = '专业培养'
peiyang_cell.alignment = align
peiyang_cell.font = Font(name="宋体 ", size=15, color='F8F8FF', bold=True, 
            italic=False, strike=False)
peiyang_cell.fill = PatternFill(fill_type='solid', start_color="4682B4", end_color='00FF00')
peiyang_cell.border=Border(top=side,right=side)
ws.merge_cells(start_row=3, start_column=1, end_row=3+max1+max2, end_column=1)

#部类核心课
buleihexin_title = ws.cell(row=3, column=2)
buleihexin_title.value = '部类核心课'
buleihexin_title.alignment = align
buleihexin_title.fill = PatternFill(fill_type='solid', start_color="87CEFA", end_color='00FF00')
buleihexin_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
buleihexin_title.border = Border(top=side,right=side)
ws.merge_cells(start_row=3, start_column=2, end_row=3+max1-1, end_column=2)

for i in range(len(ls_buleihexin)):
    for j in range(len(ls_buleihexin[i])):
        course_cell = ws.cell(row=3+j, column=3+i)
        course_cell.value = ls_buleihexin[i][j]
        course_cell.alignment = align
        if course_cell.value != '':
            course_cell.font = Font(name="宋体 ", size=12, color='000000', bold=False, 
                italic=False, strike=False)
            course_cell.fill = PatternFill(fill_type='solid', start_color="F0F8FF", end_color='00FF00')
            course_cell.border = Border(top=side,left=side)

#专业核心课
zhuanyehexin_title = ws.cell(row=3+max1, column=2)
zhuanyehexin_title.value = '专业核心课'
zhuanyehexin_title.alignment = align
zhuanyehexin_title.fill = PatternFill(fill_type='solid', start_color="87CEEB", end_color='00FF00')
zhuanyehexin_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
zhuanyehexin_title.border = Border(top=side,right=side)
ws.merge_cells(start_row=3+max1, start_column=2, end_row=3+max1+max2-1, end_column=2)

for i in range(len(ls_zhuanyehexin)):
    for j in range(len(ls_zhuanyehexin[i])):
        course_cell = ws.cell(row=3+max1+j,column = 3+i)
        course_cell.value = ls_zhuanyehexin[i][j]
        course_cell.alignment = align
        if course_cell.value != '':
            course_cell.font = Font(name="宋体 ", size=12, color='000000', bold=False, 
                italic=False, strike=False)
            course_cell.fill = PatternFill(fill_type='solid', start_color="F0F8FF", end_color='00FF00')
            course_cell.border = Border(top=side,left=side)


#个性化选修
xuanxiu_title = ws.cell(row=3+max1+max2, column=2)
xuanxiu_title.value = '个性化选修'
xuanxiu_title.alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, 
                  wrap_text=True, shrink_to_fit=False, indent=0)
xuanxiu_title.fill = PatternFill(fill_type='solid', start_color="00BFFF", end_color='00FF00')
xuanxiu_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
xuanxiu_title.border = Border(top=side,right=side)

xuanxiu_content = ws.cell(row=3+max1+max2, column=4)
xuanxiu_content.value = '计算机类、人文类、法政与社会类、管理类、理工类、经济类'
xuanxiu_content.alignment = align
xuanxiu_content.fill = PatternFill(fill_type='solid', start_color="F0F8FF", end_color='00FF00')
xuanxiu_content.font = Font(name="宋体 ", size=12, color='4169E1', bold=True, 
            italic=False, strike=False)
xuanxiu_content.border = Border(top=side)
ws.merge_cells(start_row=3+max1+max2, start_column=4, end_row=3+max1+max2, end_column=10)


#通识培养
row_tongshi = 3+max1+max2+1
tongshi_title = ws.cell(row=row_tongshi, column=1)
tongshi_title.value = '通识培养'
tongshi_title.alignment = align
tongshi_title.fill = PatternFill(fill_type='solid', start_color="FF8C00", end_color='00FF00')
tongshi_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
tongshi_title.border = Border(top=side,right=side)
ws.merge_cells(start_row=row_tongshi, start_column=1, end_row=row_tongshi, end_column=2)

tongshi_content = ws.cell(row=row_tongshi, column=3)
tongshi_content.value = '思想政治理论课、基础技能、通识课程群、公共体育、国际小学期全英文课'
tongshi_content.alignment = align
tongshi_content.fill = PatternFill(fill_type='solid', start_color="FFF8DC", end_color='00FF00')
tongshi_content.font = Font(name="宋体 ", size=12, color='DAA520', bold=True, 
            italic=False, strike=False)
tongshi_content.border = Border(top=side)
ws.merge_cells(start_row=row_tongshi, start_column=3, end_row=row_tongshi, end_column=10)


#创新训练与科学研究
row_chuangxin = 3+max1+max2+2
chuangxin_title = ws.cell(row=row_chuangxin, column=1) 
chuangxin_title.value = '创新训练与科学研究'
chuangxin_title.alignment = align
chuangxin_title.fill = PatternFill(fill_type='solid', start_color="5F9EA0", end_color='00FF00')
chuangxin_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
chuangxin_title.border = Border(top=side,right=side)
ws.merge_cells(start_row=row_chuangxin, start_column=1, end_row=row_chuangxin, end_column=2)

chuangxin_content = ws.cell(row=row_chuangxin, column=3)
chuangxin_content.value = '研究训练、其他专业实践活动'
chuangxin_content.alignment = align
chuangxin_content.fill = PatternFill(fill_type='solid', start_color="AFEEEE", end_color='00FF00')
chuangxin_content.font = Font(name="宋体 ", size=12, color='008080', bold=True, 
            italic=False, strike=False)
chuangxin_content.border = Border(top=side)
ws.merge_cells(start_row=row_chuangxin, start_column=3, end_row=row_chuangxin, end_column=10)

#素质拓展与发展指导
row_sutuo = 3+max1+max2+3
sutuo_title = ws.cell(row=row_sutuo, column=1)
sutuo_title.value = '素质拓展与发展指导'
sutuo_title.alignment = align
sutuo_title.fill = PatternFill(fill_type='solid', start_color="7B68EE", end_color='00FF00')
sutuo_title.font = Font(name="宋体 ", size=12, color='F8F8FF', bold=True, 
            italic=False, strike=False)
sutuo_title.border = Border(top=side,right=side)
ws.merge_cells(start_row=row_sutuo, start_column=1, end_row=row_sutuo, end_column=2)

sutuo_content = ws.cell(row=row_sutuo, column=3)
sutuo_content.value = '公共选修课、劳动教育、军事课、职业生涯规划、志愿服务'
sutuo_content.alignment = align
sutuo_content.fill = PatternFill(fill_type='solid', start_color="E6E6FA", end_color='00FF00')
sutuo_content.font = Font(name="宋体 ", size=12, color='483D8B', bold=True, 
            italic=False, strike=False)
sutuo_content.border = Border(top=side)

ws.merge_cells(start_row=row_sutuo, start_column=3, end_row=row_sutuo, end_column=10)
# 处理完成后保存表格，会在当前目录生成一个excel文件
wb.save(filename='cell.xlsx')
# 关闭表格对象
wb.close()
